#!/usr/bin/env python3

import openpyxl
import sys
from os import path, environ

columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']

countries = [
    9,
    300,
    301,
    302,
    303,
    304,
    305,
    306,
    307,
    308,
    311,
    312,
    313,
    314,
    315,
    316
]


def filter(sh):
    removed = False
    n_del = 0

    while not removed:
        deleted = False
        for row in range(2, sh.max_row):
            c = sh['F' + str(row)].value
            if c not in countries:
                sh.delete_rows(row)
                deleted = True
                n_del += 1
                break

        removed = not deleted

    print(f'deleted {n_del} rows')


def dedupe(sh):
    removed = False
    dupes = 0

    while not removed:
        deleted = False
        prev = None
        for row in range(2, sh.max_row):
            content = '_'.join([str(sh[col + str(row)].value) for col in columns])
            if content == prev:
                sh.delete_rows(row)
                deleted = True
                dupes += 1
                break
            prev = content

        removed = not deleted

    print(f'removed {dupes} duplicates')


def main(argv):
    action = argv[1]
    infile = argv[2]
    outfile = argv[3] if len(argv) > 3 else infile

    root = path.join(environ['HOME'], 'pr/data/xl')

    wb = openpyxl.load_workbook(path.join(root, infile + '.xlsx'))
    sh = wb.get_active_sheet()

    if action == 'filter':
        filter(sh)
    elif action == 'dedupe':
        dedupe(sh)
    else:
        print('Unsupported action')
        exit(1)

    wb.save(path.join(root, outfile + 'xlsx'))


if __name__ == '__main__':
    main(sys.argv)